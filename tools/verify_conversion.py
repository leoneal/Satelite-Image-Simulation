"""Step 1: Verify osculating Keplerian -> ECI Cartesian conversion against distance column."""
import openpyxl
import math

MU = 3.986004418e14  # m^3/s^2, Earth gravitational parameter

def kepler2cart(a, e, inc, raan, argp, nu):
    """Osculating elements -> ECI position/velocity. a in meters, angles in radians."""
    p = a * (1.0 - e * e)
    if p <= 0:
        return None, None
    # Perifocal coordinates
    denom = 1.0 + e * math.cos(nu)
    if abs(denom) < 1e-12:
        return None, None
    r_pf = [p * math.cos(nu) / denom, p * math.sin(nu) / denom, 0.0]
    h = math.sqrt(MU * p)
    v_pf = [-h / p * math.sin(nu), h / p * (e + math.cos(nu)), 0.0]

    # Rotation: R = Rz(-RAAN) Rx(-i) Rz(-argp)
    co, so = math.cos(raan), math.sin(raan)
    ci, si = math.cos(inc), math.sin(inc)
    cw, sw = math.cos(argp), math.sin(argp)

    # Combined rotation matrix (perifocal -> ECI)
    R11 = co * cw - so * sw * ci
    R12 = -co * sw - so * cw * ci
    R21 = so * cw + co * sw * ci
    R22 = -so * sw + co * cw * ci
    R31 = sw * si
    R32 = cw * si

    r = [R11 * r_pf[0] + R12 * r_pf[1],
         R21 * r_pf[0] + R22 * r_pf[1],
         R31 * r_pf[0] + R32 * r_pf[1]]
    v = [R11 * v_pf[0] + R12 * v_pf[1],
         R21 * v_pf[0] + R22 * v_pf[1],
         R31 * v_pf[0] + R32 * v_pf[1]]
    return r, v

def dist(a, b):
    return math.sqrt(sum((x - y) ** 2 for x, y in zip(a, b)))

wb = openpyxl.load_workbook(r'F:\钱室\卫星图像仿真\给钱室（公开）.xlsx', read_only=True, data_only=True)
ws = wb['Sheet1']

# Sample rows: start, a few mid, closest approach, end
check_rows = [2, 3, 4, 1000, 10000, 23388, 23389, 23390, 30000, 34202]
data = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    pass  # need index; simpler: collect all then sample

wb.close()
wb = openpyxl.load_workbook(r'F:\钱室\卫星图像仿真\给钱室（公开）.xlsx', read_only=True, data_only=True)
ws = wb['Sheet1']
all_rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

print(f"Total data rows: {len(all_rows)}")
print(f"{'row':>6} {'self_r(km)':>12} {'tar_r(km)':>12} {'calc_d(km)':>12} {'excel_d(km)':>12} {'err(km)':>10} {'cont(km)':>10}")

prev_self = None
errors = []
for idx in [r - 2 for r in check_rows if r - 2 < len(all_rows)]:
    row = all_rows[idx]
    sa, se, si, som, sw, sf = row[0], row[1], row[2], row[3], row[4], row[5]
    ta, te, ti, tom, tw, tf = row[6], row[7], row[8], row[9], row[10], row[11]
    d_excel = row[12]

    rs, vs = kepler2cart(sa, se, si, som, sw, sf)
    rt, vt = kepler2cart(ta, te, ti, tom, tw, tf)
    if rs is None or rt is None:
        print(f"row {idx+2}: CONVERSION FAILED")
        continue

    r_self_km = math.sqrt(sum(x * x for x in rs)) / 1000
    r_tar_km = math.sqrt(sum(x * x for x in rt)) / 1000
    d_calc = dist(rs, rt) / 1000
    err = abs(d_calc - d_excel)
    errors.append(err)

    cont = ""
    if prev_self is not None:
        cont = f"{dist(rs, prev_self)/1000:10.3f}"
    prev_self = rs

    print(f"{idx+2:>6} {r_self_km:12.1f} {r_tar_km:12.1f} {d_calc:12.2f} {d_excel:12.2f} {err:10.3f} {cont:>10}")

print()
max_err = max(errors)
print(f"Max distance error: {max_err:.3f} km")
if max_err < 1.0:
    print("PASS: conversion is correct (assumption: omega=RAAN, w=arg perigee, f=true anomaly, angles in rad, a in m)")
else:
    print("FAIL: need to check column meanings")
